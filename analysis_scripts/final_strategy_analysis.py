#!/usr/bin/env python3
"""FINAL STRATEGY ANALYSIS — The definitive V75/V85 profitability study.

Budget: 3000-4000 kr per round.
Key innovation: PARTIAL gardering (not full oppen, not just top-3).
Uses super_score = 0.5 * model + 0.5 * market.
Tests 20+ confidence formulas, selective play filters, dynamic budget,
estimated payouts, and expert signal comparison.

Run: python3 final_strategy_analysis.py
"""

from __future__ import annotations

import asyncio
import itertools
import json
import math
import statistics
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).parent))

from trav_agent.analysis.composite import CompositeAnalyzer
from trav_agent.config import AnalysisConfig
from trav_agent.data.atg_client import ATGClient
from trav_agent.data.models import GameRound

import logging

logging.basicConfig(
    level=logging.WARNING, format="%(asctime)s  %(message)s", datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# ── Constants ────────────────────────────────────────────────────────────────

ROW_PRICE = {"V75": 0.50, "V85": 1.00}
PRIZE_POOL_SHARE = 0.35
DEFAULT_POOL = {"V75": 30_000_000, "V85": 5_000_000}

# V75 prize: ~10.5M to 7-rätt, V85: ~1.75M to 8-rätt
V75_PRIZE_7RATT = 10_500_000
V85_PRIZE_8RATT = 1_750_000
V75_POOL_ROWS = 60_000_000
V85_POOL_ROWS = 5_000_000


def product(iterable):
    result = 1
    for x in iterable:
        result *= x
    return result


def filter_future_starts(game_round: GameRound) -> None:
    for race in game_round.races:
        for entry in race.entries:
            entry.horse.past_starts = [
                s for s in entry.horse.past_starts if s.start_date < race.race_date
            ]


def sep(title, char="=", width=110):
    print(f"\n{char * width}")
    print(f"  {title}")
    print(f"{char * width}")


# ── Data structures ──────────────────────────────────────────────────────────

@dataclass
class EntryInfo:
    post_position: int
    composite_score: float
    bet_percentage: float  # streckprocent as decimal (0.25 = 25%)
    super_score: float = 0.0
    horse_name: str = ""
    recommendation: str = ""
    factor_scores: dict = field(default_factory=dict)


@dataclass
class RaceInfo:
    race_number: int
    track_name: str
    num_starters: int
    start_method: str
    distance: int
    winner_pp: int
    entries: list[EntryInfo]  # sorted by composite_score desc initially
    upset_risk: float
    gap_to_second: float
    gap_to_third: float


@dataclass
class RoundInfo:
    game_type: str
    game_id: str
    round_date: date
    track_name: str
    turnover: Optional[int]
    jackpot: Optional[int]
    races: list[RaceInfo]


# ── Data Loading ─────────────────────────────────────────────────────────────

async def load_rounds() -> list[GameRound]:
    client = ATGClient()
    all_rounds: list[GameRound] = []
    end = date(2026, 2, 21)

    for gt, start in [("V75", date(2024, 1, 1)), ("V85", date(2024, 3, 1))]:
        logger.info(f"Loading {gt}...")
        async for day, gr in client.fetch_historical_rounds_iter(gt, start, end):
            if gr and gr.is_finished:
                all_rounds.append(gr)

    logger.info(f"Total loaded: {len(all_rounds)} rounds")
    return all_rounds


def analyze_all_rounds(all_rounds: list[GameRound]) -> list[RoundInfo]:
    round_infos = []
    skipped = 0

    for idx, gr in enumerate(all_rounds):
        gr_copy = gr.model_copy(deep=True)
        filter_future_starts(gr_copy)

        analyzer = CompositeAnalyzer(AnalysisConfig())
        analyzer.analyze_round(gr_copy)

        expected_races = 7 if gr.game_type == "V75" else 8

        valid = True
        for race in gr_copy.races:
            if not race.result_order or not race.active_entries:
                valid = False
                break
        if not valid or len(gr_copy.races) < expected_races:
            skipped += 1
            continue

        ri = RoundInfo(
            game_type=gr.game_type,
            game_id=gr.game_id,
            round_date=gr.round_date,
            track_name=gr.track_name,
            turnover=gr.turnover,
            jackpot=gr.jackpot,
            races=[],
        )

        for race in gr_copy.races:
            winner_pp = race.result_order[0]

            sorted_entries = sorted(
                race.active_entries,
                key=lambda e: e.composite_score,
                reverse=True,
            )

            winner_found = any(e.post_position == winner_pp for e in sorted_entries)
            if not winner_found:
                valid = False
                break

            entries = []
            for e in sorted_entries:
                entries.append(EntryInfo(
                    post_position=e.post_position,
                    composite_score=e.composite_score,
                    bet_percentage=e.bet_percentage or 0.0,
                    horse_name=e.horse.name,
                    recommendation=e.recommendation,
                    factor_scores=dict(e.factor_scores),
                ))

            gap2 = 0.0
            gap3 = 0.0
            if len(sorted_entries) >= 2:
                gap2 = sorted_entries[0].composite_score - sorted_entries[1].composite_score
            if len(sorted_entries) >= 3:
                gap3 = sorted_entries[0].composite_score - sorted_entries[2].composite_score

            race_info = RaceInfo(
                race_number=race.race_number,
                track_name=race.track_name,
                num_starters=race.num_starters,
                start_method=race.start_method.value,
                distance=race.distance,
                winner_pp=winner_pp,
                entries=entries,
                upset_risk=race.upset_risk,
                gap_to_second=gap2,
                gap_to_third=gap3,
            )
            ri.races.append(race_info)

        if not valid or len(ri.races) < expected_races:
            skipped += 1
            continue

        round_infos.append(ri)

        if (idx + 1) % 20 == 0:
            logger.info(f"  Analyzed {idx+1}/{len(all_rounds)}...")

    logger.info(f"Total valid rounds: {len(round_infos)} (skipped {skipped})")
    return round_infos


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1: COMPUTE SUPER SCORES
# ══════════════════════════════════════════════════════════════════════════════

def compute_super_scores_for_race(race: RaceInfo, alpha: float = 0.5) -> list[EntryInfo]:
    """Compute super_score = (1-alpha)*composite + alpha*market_normalized.

    Normalizes market (streck) within the field to similar scale as composite.
    Returns entries sorted by super_score descending.
    """
    composites = [e.composite_score for e in race.entries]
    markets = [e.bet_percentage for e in race.entries]

    min_m = min(markets) if markets else 0
    max_m = max(markets) if markets else 1
    spread = max_m - min_m if max_m > min_m else 1.0

    for e in race.entries:
        market_norm = ((e.bet_percentage - min_m) / spread) * 80 + 10  # scale to 10-90
        e.super_score = e.composite_score * (1.0 - alpha) + market_norm * alpha

    return sorted(race.entries, key=lambda e: e.super_score, reverse=True)


def compute_all_super_scores(round_infos: list[RoundInfo], alpha: float = 0.5):
    """Compute super_scores for all races in all rounds. Modifies entries in place."""
    for ri in round_infos:
        for race in ri.races:
            compute_super_scores_for_race(race, alpha)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2: CONFIDENCE METRICS — 20+ formulas
# ══════════════════════════════════════════════════════════════════════════════

def get_race_stats(race: RaceInfo) -> dict:
    """Compute various statistics for a race needed by confidence formulas."""
    sorted_by_super = sorted(race.entries, key=lambda e: e.super_score, reverse=True)
    sorted_by_market = sorted(race.entries, key=lambda e: e.bet_percentage, reverse=True)
    sorted_by_model = sorted(race.entries, key=lambda e: e.composite_score, reverse=True)

    s1 = sorted_by_super[0].super_score if sorted_by_super else 0
    s2 = sorted_by_super[1].super_score if len(sorted_by_super) >= 2 else 0
    s3 = sorted_by_super[2].super_score if len(sorted_by_super) >= 3 else 0
    avg_super = statistics.mean([e.super_score for e in race.entries]) if race.entries else 0

    m1 = sorted_by_market[0].bet_percentage if sorted_by_market else 0
    m2 = sorted_by_market[1].bet_percentage if len(sorted_by_market) >= 2 else 0

    c1 = sorted_by_model[0].composite_score if sorted_by_model else 0
    avg_composite = statistics.mean([e.composite_score for e in race.entries]) if race.entries else 0

    # Does model #1 agree with market #1?
    model_top_pp = sorted_by_model[0].post_position if sorted_by_model else -1
    market_top_pp = sorted_by_market[0].post_position if sorted_by_market else -2
    super_top_pp = sorted_by_super[0].post_position if sorted_by_super else -3
    agreement = model_top_pp == market_top_pp

    # streckprocent std (spread)
    strecks = [e.bet_percentage for e in race.entries if e.bet_percentage > 0]
    streck_std = statistics.stdev(strecks) if len(strecks) >= 2 else 0

    return {
        "s1": s1, "s2": s2, "s3": s3, "avg_super": avg_super,
        "super_gap_1_2": s1 - s2,
        "super_gap_1_3": s1 - s3,
        "m1": m1, "m2": m2,
        "market_gap": m1 - m2,
        "c1": c1, "avg_composite": avg_composite,
        "model_dominance": c1 / avg_composite if avg_composite > 0 else 1,
        "agreement": agreement,
        "model_top_pp": model_top_pp,
        "market_top_pp": market_top_pp,
        "super_top_pp": super_top_pp,
        "streck_std": streck_std,
        "num_starters": race.num_starters,
        "start_method": race.start_method,
        "upset_risk": race.upset_risk,
        "gap_to_second_composite": race.gap_to_second,
        "sorted_by_super": sorted_by_super,
    }


def define_confidence_formulas() -> dict:
    """Define 25+ confidence formulas. Each returns higher = more confident."""
    formulas = {}

    # A) Simple super_score gap to 2nd
    formulas["A_super_gap"] = lambda s: s["super_gap_1_2"]

    # B) Super gap * market leader streck
    formulas["B_gap_x_streck"] = lambda s: s["super_gap_1_2"] * s["m1"] * 100

    # C) Super gap to 3rd / num_starters
    formulas["C_gap3_div_starters"] = lambda s: s["super_gap_1_3"] / max(s["num_starters"], 1) * 10

    # D) Market confidence: gap between #1 and #2 streck
    formulas["D_market_gap"] = lambda s: s["market_gap"] * 100

    # E) Model dominance: composite #1 / avg
    formulas["E_model_dominance"] = lambda s: (s["model_dominance"] - 1) * 100

    # F) Super score gap (same as A but scaled)
    formulas["F_super_gap_scaled"] = lambda s: s["super_gap_1_2"] * 2

    # G) Gap * auto bonus (penalize volt)
    formulas["G_gap_auto"] = lambda s: s["super_gap_1_2"] * (1.0 if s["start_method"] == "auto" else 0.7)

    # H) Agreement bonus + gap
    formulas["H_agreement_gap"] = lambda s: s["super_gap_1_2"] + (10 if s["agreement"] else 0)

    # I) Market streck #1 alone
    formulas["I_streck_1st"] = lambda s: s["m1"] * 100

    # J) Combined: gap * streck * agreement
    formulas["J_combined"] = lambda s: (s["super_gap_1_2"] * s["m1"] * 100 *
                                         (1.3 if s["agreement"] else 0.8))

    # K) Gap / (1 + upset_risk/100)
    formulas["K_gap_div_upset"] = lambda s: s["super_gap_1_2"] / (1 + s["upset_risk"] / 100)

    # L) (streck_1 - streck_2) * gap
    formulas["L_streck_gap_product"] = lambda s: s["market_gap"] * s["super_gap_1_2"] * 100

    # M) Gap * (1 + streck_1) / (1 + starters/15)
    formulas["M_original"] = lambda s: (s["super_gap_1_2"] *
                                         (1 + s["m1"]) /
                                         (1 + s["num_starters"] / 15))

    # N) Streck std (high std = one horse dominates = confident)
    formulas["N_streck_std"] = lambda s: s["streck_std"] * 100

    # O) Combined: gap + streck*30 + agreement*5 - starters
    formulas["O_multifeature"] = lambda s: (s["super_gap_1_2"] +
                                             s["m1"] * 30 +
                                             (5 if s["agreement"] else 0) -
                                             s["num_starters"] * 0.5)

    # P) Gap^2 (rewards big gaps more)
    formulas["P_gap_squared"] = lambda s: s["super_gap_1_2"] ** 2 / 10

    # Q) Model dominance * market gap
    formulas["Q_dom_x_mktgap"] = lambda s: s["model_dominance"] * s["market_gap"] * 100

    # R) Inverse upset risk
    formulas["R_inv_upset"] = lambda s: 100 - s["upset_risk"]

    # S) Gap * (1 - upset_risk/100)
    formulas["S_gap_x_safety"] = lambda s: s["super_gap_1_2"] * (1 - s["upset_risk"] / 100)

    # T) Composite gap (original model gap, not super)
    formulas["T_composite_gap"] = lambda s: s["gap_to_second_composite"]

    # U) Best combined: gap * streck * (1 if auto else 0.8) * agreement_bonus / starters
    formulas["U_kitchen_sink"] = lambda s: (
        s["super_gap_1_2"] *
        (s["m1"] * 100 + 5) *
        (1.0 if s["start_method"] == "auto" else 0.8) *
        (1.2 if s["agreement"] else 0.9) /
        max(s["num_starters"], 5)
    )

    # V) Super gap * market gap combined
    formulas["V_dual_gap"] = lambda s: s["super_gap_1_2"] * s["market_gap"] * 1000

    # W) Simple: streck_1 > 0.25 is confident
    formulas["W_favorite_binary"] = lambda s: 50 if s["m1"] > 0.25 else (25 if s["m1"] > 0.15 else 5)

    # X) Weighted sum: 0.3*gap + 0.3*streck + 0.2*agreement + 0.2*inv_starters
    formulas["X_weighted_sum"] = lambda s: (
        s["super_gap_1_2"] * 3 +
        s["m1"] * 30 +
        (10 if s["agreement"] else 0) +
        max(0, 15 - s["num_starters"])
    )

    # Y) Entropy-based: low entropy in streck = confident
    formulas["Y_low_entropy"] = lambda s: (s["m1"] - s["m2"]) * 100 + s["super_gap_1_2"]

    return formulas


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3: PARTIAL GARDERING SIMULATION
# ══════════════════════════════════════════════════════════════════════════════

def assign_picks_for_round(ri: RoundInfo, race_confidences: list[float],
                            max_rows: int) -> list[int]:
    """Assign number of picks per race using confidence-based partial gardering.

    Algorithm:
    1. Start with 1 pick per race
    2. Iteratively add 1 pick to the LEAST confident race that has room
    3. Stop when total_rows would exceed max_rows
    """
    n_races = len(ri.races)
    picks = [1] * n_races

    # Sort race indices by confidence (most confident first)
    indices_by_conf = sorted(range(n_races), key=lambda i: race_confidences[i], reverse=True)

    while True:
        current_rows = product(picks)
        if current_rows >= max_rows:
            break

        expanded = False
        # Try to expand least confident race first
        for idx in reversed(indices_by_conf):
            race = ri.races[idx]
            if picks[idx] >= race.num_starters:
                continue

            new_picks = picks[:]
            new_picks[idx] += 1
            new_rows = product(new_picks)

            if new_rows <= max_rows:
                picks[idx] += 1
                expanded = True
                break

        if not expanded:
            break

    return picks


def simulate_round(ri: RoundInfo, race_confidences: list[float],
                    max_rows: int) -> dict:
    """Simulate a round with partial gardering strategy.

    Returns picks, cost, hit status, payout info.
    """
    n_races = len(ri.races)
    picks_counts = assign_picks_for_round(ri, race_confidences, max_rows)

    total_rows = product(picks_counts)
    row_price = ROW_PRICE.get(ri.game_type, 1.0)
    cost = total_rows * row_price

    # Check hit: is winner in top-N picks for each race?
    hit = True
    winner_strecks = []
    picks_detail = []

    for i, race in enumerate(ri.races):
        sorted_by_super = sorted(race.entries, key=lambda e: e.super_score, reverse=True)
        our_picks = [e.post_position for e in sorted_by_super[:picks_counts[i]]]
        picks_detail.append(our_picks)

        # Find winner streck
        winner_entry = next((e for e in race.entries if e.post_position == race.winner_pp), None)
        winner_streck = winner_entry.bet_percentage if winner_entry else 0.001
        winner_strecks.append(max(winner_streck, 0.001))

        if race.winner_pp not in our_picks:
            hit = False

    # Estimate payout
    streck_product = product(winner_strecks)
    pool = ri.turnover if ri.turnover else DEFAULT_POOL.get(ri.game_type, 10_000_000)
    prize_pool = pool * PRIZE_POOL_SHARE
    pool_rows = pool / row_price
    expected_winners = pool_rows * streck_product
    if expected_winners < 0.001:
        expected_winners = 0.001
    est_payout = min(prize_pool, prize_pool / expected_winners)

    # Also compute payout using fixed prize pools
    if ri.game_type == "V75":
        fixed_prize = V75_PRIZE_7RATT
        fixed_pool_rows = V75_POOL_ROWS
    else:
        fixed_prize = V85_PRIZE_8RATT
        fixed_pool_rows = V85_POOL_ROWS
    fixed_expected_winners = fixed_pool_rows * streck_product
    if fixed_expected_winners < 0.001:
        fixed_expected_winners = 0.001
    fixed_payout = min(fixed_prize, fixed_prize / fixed_expected_winners)

    return {
        "picks_counts": picks_counts,
        "picks_detail": picks_detail,
        "total_rows": total_rows,
        "cost": cost,
        "hit": hit,
        "est_payout": est_payout,
        "fixed_payout": fixed_payout,
        "winner_strecks": winner_strecks,
        "streck_product": streck_product,
    }


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4: MAIN ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

def run_full_analysis(round_infos: list[RoundInfo]) -> dict:
    """Run the complete analysis pipeline."""
    results = {}

    sep("STEP 1: SUPER SCORE COMPUTATION")
    compute_all_super_scores(round_infos, alpha=0.5)

    # Verify super_score accuracy
    total_races = sum(len(ri.races) for ri in round_infos)
    top1_hits = 0
    top2_hits = 0
    top3_hits = 0

    for ri in round_infos:
        for race in ri.races:
            sorted_by_super = sorted(race.entries, key=lambda e: e.super_score, reverse=True)
            ranking = [e.post_position for e in sorted_by_super]
            if race.winner_pp in ranking:
                w_rank = ranking.index(race.winner_pp) + 1
            else:
                w_rank = len(ranking) + 1
            if w_rank == 1: top1_hits += 1
            if w_rank <= 2: top2_hits += 1
            if w_rank <= 3: top3_hits += 1

    print(f"\n  Super score (alpha=0.5) accuracy on {total_races} races:")
    print(f"    Top-1: {top1_hits}/{total_races} = {top1_hits/total_races:.1%}")
    print(f"    Top-2: {top2_hits}/{total_races} = {top2_hits/total_races:.1%}")
    print(f"    Top-3: {top3_hits}/{total_races} = {top3_hits/total_races:.1%}")

    results["super_score_accuracy"] = {
        "top1": top1_hits / total_races,
        "top2": top2_hits / total_races,
        "top3": top3_hits / total_races,
    }

    # ══════════════════════════════════════════════════════════════════════
    # STEP 2: TEST 25 CONFIDENCE FORMULAS
    # ══════════════════════════════════════════════════════════════════════

    sep("STEP 2: TEST 25 CONFIDENCE FORMULAS")

    confidence_formulas = define_confidence_formulas()
    budgets_to_test = [3000, 3500, 4000]

    formula_results = {}

    print(f"\n  Testing {len(confidence_formulas)} confidence formulas x {len(budgets_to_test)} budgets")
    print(f"  across {len(round_infos)} rounds")
    print(f"\n  {'Formula':<25} {'Budget':>7} {'Hits':>5} {'HitRate':>8} {'Cost':>10} {'Payout':>12} {'ROI':>8} {'Netto':>12}")
    print(f"  {'-'*25} {'-'*7} {'-'*5} {'-'*8} {'-'*10} {'-'*12} {'-'*8} {'-'*12}")

    best_overall = {"formula": "", "budget": 0, "roi": -999, "data": {}}

    for formula_name, formula_fn in confidence_formulas.items():
        formula_results[formula_name] = {}

        for budget in budgets_to_test:
            total_cost = 0
            total_payout = 0
            hits = 0
            hit_details = []

            for ri in round_infos:
                row_price = ROW_PRICE.get(ri.game_type, 1.0)
                max_rows = int(budget / row_price)

                # Compute confidence per race
                race_confs = []
                for race in ri.races:
                    stats = get_race_stats(race)
                    try:
                        conf = formula_fn(stats)
                    except Exception:
                        conf = 0
                    race_confs.append(conf)

                sim = simulate_round(ri, race_confs, max_rows)
                total_cost += sim["cost"]

                if sim["hit"]:
                    hits += 1
                    total_payout += sim["fixed_payout"]
                    hit_details.append({
                        "date": ri.round_date.isoformat(),
                        "game_type": ri.game_type,
                        "cost": sim["cost"],
                        "payout": sim["fixed_payout"],
                        "picks": sim["picks_counts"],
                        "rows": sim["total_rows"],
                    })

            roi = (total_payout - total_cost) / total_cost if total_cost > 0 else 0
            netto = total_payout - total_cost
            hit_rate = hits / len(round_infos)

            formula_results[formula_name][budget] = {
                "hits": hits,
                "hit_rate": hit_rate,
                "total_cost": total_cost,
                "total_payout": total_payout,
                "roi": roi,
                "netto": netto,
                "hit_details": hit_details,
            }

            if roi > best_overall["roi"]:
                best_overall = {
                    "formula": formula_name,
                    "budget": budget,
                    "roi": roi,
                    "data": formula_results[formula_name][budget],
                }

            print(f"  {formula_name:<25} {budget:>7} {hits:>5} {hit_rate:>7.1%} "
                  f"{total_cost:>10,.0f} {total_payout:>12,.0f} {roi*100:>+7.1f}% {netto:>+12,.0f}")

    # Show best formulas
    sep("TOP 10 CONFIDENCE FORMULAS (by ROI at best budget)")

    all_formula_roi = []
    for fname, budgets_data in formula_results.items():
        best_budget_roi = max(budgets_data.items(), key=lambda x: x[1]["roi"])
        all_formula_roi.append((fname, best_budget_roi[0], best_budget_roi[1]))

    all_formula_roi.sort(key=lambda x: x[2]["roi"], reverse=True)

    print(f"\n  {'#':>3} {'Formula':<25} {'Budget':>7} {'Hits':>5} {'ROI':>8} {'Netto':>12}")
    print(f"  {'-'*3} {'-'*25} {'-'*7} {'-'*5} {'-'*8} {'-'*12}")

    for i, (fname, budget, data) in enumerate(all_formula_roi[:10]):
        print(f"  {i+1:>3} {fname:<25} {budget:>7} {data['hits']:>5} "
              f"{data['roi']*100:>+7.1f}% {data['netto']:>+12,.0f}")

    results["best_formula"] = best_overall
    results["formula_ranking"] = [
        {"formula": f, "budget": b, "roi": d["roi"], "hits": d["hits"], "netto": d["netto"]}
        for f, b, d in all_formula_roi
    ]

    # Use best formula for subsequent analysis
    best_formula_name = best_overall["formula"]
    best_budget = best_overall["budget"]
    best_formula_fn = confidence_formulas[best_formula_name]

    print(f"\n  SELECTED: {best_formula_name} @ {best_budget} kr")

    # ══════════════════════════════════════════════════════════════════════
    # STEP 3: SELECTIVE PLAY FILTERS
    # ══════════════════════════════════════════════════════════════════════

    sep("STEP 3: SELECTIVE PLAY FILTERS")

    # Pre-compute race stats for all rounds
    round_race_stats = []
    for ri in round_infos:
        race_stats = []
        for race in ri.races:
            stats = get_race_stats(race)
            race_stats.append(stats)
        round_race_stats.append(race_stats)

    # Pre-compute round-level confidence metrics
    round_conf_metrics = []
    for ri_idx, ri in enumerate(round_infos):
        confs = []
        for rs in round_race_stats[ri_idx]:
            try:
                c = best_formula_fn(rs)
            except:
                c = 0
            confs.append(c)

        n_agreement = sum(1 for rs in round_race_stats[ri_idx] if rs["agreement"])
        avg_conf = statistics.mean(confs) if confs else 0
        min_conf = min(confs) if confs else 0
        avg_starters = statistics.mean([r.num_starters for r in ri.races])
        n_volt = sum(1 for r in ri.races if r.start_method == "volt")
        avg_streck_1 = statistics.mean([rs["m1"] for rs in round_race_stats[ri_idx]])
        n_big_field = sum(1 for r in ri.races if r.num_starters > 12)
        avg_upset = statistics.mean([r.upset_risk for r in ri.races])

        # Payout estimation (pre-hit)
        # Use average favorite streck as proxy for expected difficulty
        expected_fav_product = product([max(rs["m1"], 0.05) for rs in round_race_stats[ri_idx]])
        if ri.game_type == "V75":
            est_payout_if_fav = V75_PRIZE_7RATT / max(V75_POOL_ROWS * expected_fav_product, 0.001)
        else:
            est_payout_if_fav = V85_PRIZE_8RATT / max(V85_POOL_ROWS * expected_fav_product, 0.001)

        round_conf_metrics.append({
            "confidences": confs,
            "avg_conf": avg_conf,
            "min_conf": min_conf,
            "n_agreement": n_agreement,
            "avg_starters": avg_starters,
            "n_volt": n_volt,
            "avg_streck_1": avg_streck_1,
            "n_big_field": n_big_field,
            "avg_upset": avg_upset,
            "est_payout_if_fav": est_payout_if_fav,
        })

    # Define filters
    n_races_per = lambda ri: 7 if ri.game_type == "V75" else 8

    selective_filters = {
        "ALL": lambda ri, m: True,

        # Confidence-based
        "avg_conf_gt_median": lambda ri, m: m["avg_conf"] > statistics.median([mm["avg_conf"] for mm in round_conf_metrics]),
        "avg_conf_gt_p60": lambda ri, m: m["avg_conf"] > sorted([mm["avg_conf"] for mm in round_conf_metrics])[int(len(round_conf_metrics)*0.4)],
        "avg_conf_gt_p75": lambda ri, m: m["avg_conf"] > sorted([mm["avg_conf"] for mm in round_conf_metrics])[int(len(round_conf_metrics)*0.25)],

        # Agreement-based
        "agreement_gte_half": lambda ri, m: m["n_agreement"] >= n_races_per(ri) // 2,
        "agreement_gte_5": lambda ri, m: m["n_agreement"] >= 5,
        "agreement_gte_6": lambda ri, m: m["n_agreement"] >= 6,

        # Field size
        "no_big_field_volt": lambda ri, m: not any(r.num_starters > 12 and r.start_method == "volt" for r in ri.races),
        "max_avg_starters_10": lambda ri, m: m["avg_starters"] <= 10,
        "max_avg_starters_11": lambda ri, m: m["avg_starters"] <= 11,

        # Upset-based
        "avg_upset_lt_35": lambda ri, m: m["avg_upset"] < 35,
        "avg_upset_lt_40": lambda ri, m: m["avg_upset"] < 40,
        "no_very_high_upset": lambda ri, m: all(r.upset_risk < 70 for r in ri.races),

        # Payout-based
        "est_payout_gt_5000": lambda ri, m: m["est_payout_if_fav"] > 5000,
        "est_payout_gt_10000": lambda ri, m: m["est_payout_if_fav"] > 10000,
        "est_payout_lt_500k": lambda ri, m: m["est_payout_if_fav"] < 500000,

        # Combined
        "agreement_5_upset_40": lambda ri, m: m["n_agreement"] >= 5 and m["avg_upset"] < 40,
        "agreement_5_no_bigvolt": lambda ri, m: m["n_agreement"] >= 5 and not any(r.num_starters > 12 and r.start_method == "volt" for r in ri.races),
        "conf_med_upset_40": lambda ri, m: m["avg_conf"] > statistics.median([mm["avg_conf"] for mm in round_conf_metrics]) and m["avg_upset"] < 40,

        # Game type
        "V75_only": lambda ri, m: ri.game_type == "V75",
        "V85_only": lambda ri, m: ri.game_type == "V85",
    }

    print(f"\n  Using formula: {best_formula_name} @ {best_budget} kr budget")
    print(f"\n  {'Filter':<35} {'Played':>7} {'Hits':>5} {'Cost':>10} {'Payout':>12} {'ROI':>8} {'Netto':>12}")
    print(f"  {'-'*35} {'-'*7} {'-'*5} {'-'*10} {'-'*12} {'-'*8} {'-'*12}")

    selective_results = {}
    best_selective = {"filter": "", "roi": -999}

    for filt_name, filt_fn in selective_filters.items():
        indices = [i for i in range(len(round_infos))
                   if filt_fn(round_infos[i], round_conf_metrics[i])]

        if len(indices) < 3:
            continue

        total_cost = 0
        total_payout = 0
        hits = 0

        for idx in indices:
            ri = round_infos[idx]
            row_price = ROW_PRICE.get(ri.game_type, 1.0)
            max_rows = int(best_budget / row_price)

            confs = round_conf_metrics[idx]["confidences"]
            sim = simulate_round(ri, confs, max_rows)
            total_cost += sim["cost"]

            if sim["hit"]:
                hits += 1
                total_payout += sim["fixed_payout"]

        roi = (total_payout - total_cost) / total_cost if total_cost > 0 else 0
        netto = total_payout - total_cost

        selective_results[filt_name] = {
            "played": len(indices),
            "hits": hits,
            "total_cost": total_cost,
            "total_payout": total_payout,
            "roi": roi,
            "netto": netto,
        }

        if roi > best_selective["roi"]:
            best_selective = {"filter": filt_name, "roi": roi, "data": selective_results[filt_name]}

        print(f"  {filt_name:<35} {len(indices):>7} {hits:>5} {total_cost:>10,.0f} "
              f"{total_payout:>12,.0f} {roi*100:>+7.1f}% {netto:>+12,.0f}")

    results["selective_results"] = selective_results
    results["best_selective"] = best_selective

    # ══════════════════════════════════════════════════════════════════════
    # STEP 4: CROSS-VALIDATE: ALL FORMULAS x ALL FILTERS x BUDGETS
    # ══════════════════════════════════════════════════════════════════════

    sep("STEP 4: EXHAUSTIVE SEARCH — Formula x Filter x Budget")

    top_formulas = [f for f, _, _ in all_formula_roi[:8]]  # Top 8 formulas
    top_filters = ["ALL", "agreement_gte_5", "agreement_gte_6", "avg_upset_lt_40",
                    "no_big_field_volt", "agreement_5_upset_40", "est_payout_gt_5000",
                    "V75_only", "V85_only"]
    budgets_ext = [2000, 2500, 3000, 3500, 4000]

    exhaustive_results = []

    for formula_name in top_formulas:
        formula_fn = confidence_formulas[formula_name]

        for filt_name in top_filters:
            if filt_name not in selective_filters:
                continue
            filt_fn = selective_filters[filt_name]

            indices = [i for i in range(len(round_infos))
                       if filt_fn(round_infos[i], round_conf_metrics[i])]

            if len(indices) < 3:
                continue

            for budget in budgets_ext:
                total_cost = 0
                total_payout = 0
                hits = 0
                hit_list = []

                for idx in indices:
                    ri = round_infos[idx]
                    row_price = ROW_PRICE.get(ri.game_type, 1.0)
                    max_rows = int(budget / row_price)

                    confs = []
                    for race in ri.races:
                        stats = get_race_stats(race)
                        try:
                            c = formula_fn(stats)
                        except:
                            c = 0
                        confs.append(c)

                    sim = simulate_round(ri, confs, max_rows)
                    total_cost += sim["cost"]

                    if sim["hit"]:
                        hits += 1
                        total_payout += sim["fixed_payout"]
                        hit_list.append({
                            "date": ri.round_date.isoformat(),
                            "game_type": ri.game_type,
                            "cost": sim["cost"],
                            "payout": sim["fixed_payout"],
                            "picks": sim["picks_counts"],
                        })

                roi = (total_payout - total_cost) / total_cost if total_cost > 0 else 0
                netto = total_payout - total_cost

                exhaustive_results.append({
                    "formula": formula_name,
                    "filter": filt_name,
                    "budget": budget,
                    "played": len(indices),
                    "hits": hits,
                    "hit_rate": hits / len(indices) if indices else 0,
                    "total_cost": total_cost,
                    "total_payout": total_payout,
                    "roi": roi,
                    "netto": netto,
                    "hit_list": hit_list,
                })

    exhaustive_results.sort(key=lambda x: x["roi"], reverse=True)

    print(f"\n  Tested {len(exhaustive_results)} combinations")
    print(f"\n  TOP 30 STRATEGIES (by ROI):")
    print(f"  {'#':>3} {'Formula':<22} {'Filter':<25} {'Budget':>6} {'Played':>6} {'Hits':>4} {'HitRate':>7} {'ROI':>8} {'Netto':>12}")
    print(f"  {'-'*3} {'-'*22} {'-'*25} {'-'*6} {'-'*6} {'-'*4} {'-'*7} {'-'*8} {'-'*12}")

    for i, r in enumerate(exhaustive_results[:30]):
        print(f"  {i+1:>3} {r['formula']:<22} {r['filter']:<25} {r['budget']:>6} "
              f"{r['played']:>6} {r['hits']:>4} {r['hit_rate']:>6.1%} "
              f"{r['roi']*100:>+7.1f}% {r['netto']:>+12,.0f}")

    results["exhaustive_top30"] = exhaustive_results[:30]

    # ══════════════════════════════════════════════════════════════════════
    # STEP 5: DYNAMIC BUDGET
    # ══════════════════════════════════════════════════════════════════════

    sep("STEP 5: DYNAMIC BUDGET")

    # Use top formula
    top_formula_fn = confidence_formulas[exhaustive_results[0]["formula"]] if exhaustive_results else best_formula_fn
    top_formula_name_dyn = exhaustive_results[0]["formula"] if exhaustive_results else best_formula_name

    print(f"\n  Testing dynamic budget allocation with formula: {top_formula_name_dyn}")

    dynamic_strategies = {
        "fixed_3000": lambda ri, m: 3000,
        "fixed_3500": lambda ri, m: 3500,
        "fixed_4000": lambda ri, m: 4000,
        "high_conf_4000_else_2000": lambda ri, m: 4000 if m["avg_conf"] > statistics.median([mm["avg_conf"] for mm in round_conf_metrics]) else 2000,
        "high_payout_4000_else_2000": lambda ri, m: 4000 if m["est_payout_if_fav"] > 10000 else 2000,
        "agree_4000_else_2500": lambda ri, m: 4000 if m["n_agreement"] >= 5 else 2500,
        "upset_based": lambda ri, m: 4000 if m["avg_upset"] < 35 else (3000 if m["avg_upset"] < 45 else 2000),
        "combined_dynamic": lambda ri, m: (
            4000 if (m["avg_conf"] > statistics.median([mm["avg_conf"] for mm in round_conf_metrics]) and m["avg_upset"] < 40)
            else (3000 if m["n_agreement"] >= 4
                  else 2000)
        ),
        "skip_or_4000": lambda ri, m: 4000 if m["n_agreement"] >= 5 else 0,
        "skip_or_3500": lambda ri, m: 3500 if m["avg_upset"] < 40 else 0,
        "sniper_4000": lambda ri, m: 4000 if (m["n_agreement"] >= 5 and m["avg_upset"] < 40 and m["est_payout_if_fav"] > 5000) else 0,
    }

    print(f"\n  {'Strategy':<35} {'Played':>7} {'Hits':>5} {'Cost':>10} {'Payout':>12} {'ROI':>8} {'Netto':>12} {'AvgBet':>8}")
    print(f"  {'-'*35} {'-'*7} {'-'*5} {'-'*10} {'-'*12} {'-'*8} {'-'*12} {'-'*8}")

    dynamic_results = {}
    best_dynamic = {"strategy": "", "roi": -999}

    for strat_name, budget_fn in dynamic_strategies.items():
        total_cost = 0
        total_payout = 0
        hits = 0
        played = 0

        for ri_idx, ri in enumerate(round_infos):
            m = round_conf_metrics[ri_idx]
            budget = budget_fn(ri, m)

            if budget <= 0:
                continue  # Skip this round

            played += 1
            row_price = ROW_PRICE.get(ri.game_type, 1.0)
            max_rows = int(budget / row_price)

            confs = []
            for race in ri.races:
                stats = get_race_stats(race)
                try:
                    c = top_formula_fn(stats)
                except:
                    c = 0
                confs.append(c)

            sim = simulate_round(ri, confs, max_rows)
            total_cost += sim["cost"]

            if sim["hit"]:
                hits += 1
                total_payout += sim["fixed_payout"]

        roi = (total_payout - total_cost) / total_cost if total_cost > 0 else 0
        netto = total_payout - total_cost
        avg_bet = total_cost / played if played > 0 else 0

        dynamic_results[strat_name] = {
            "played": played,
            "hits": hits,
            "total_cost": total_cost,
            "total_payout": total_payout,
            "roi": roi,
            "netto": netto,
            "avg_bet": avg_bet,
        }

        if roi > best_dynamic["roi"]:
            best_dynamic = {"strategy": strat_name, "roi": roi, "data": dynamic_results[strat_name]}

        print(f"  {strat_name:<35} {played:>7} {hits:>5} {total_cost:>10,.0f} "
              f"{total_payout:>12,.0f} {roi*100:>+7.1f}% {netto:>+12,.0f} {avg_bet:>8,.0f}")

    results["dynamic_results"] = dynamic_results
    results["best_dynamic"] = best_dynamic

    # ══════════════════════════════════════════════════════════════════════
    # STEP 6: MONTHLY BREAKDOWN FOR BEST STRATEGY
    # ══════════════════════════════════════════════════════════════════════

    sep("STEP 6: MONTHLY BREAKDOWN")

    # Use the absolute best from exhaustive search
    if exhaustive_results and exhaustive_results[0]["roi"] > best_dynamic.get("roi", -999):
        use_formula_name = exhaustive_results[0]["formula"]
        use_formula_fn = confidence_formulas[use_formula_name]
        use_budget = exhaustive_results[0]["budget"]
        use_filter_name = exhaustive_results[0]["filter"]
        use_filter_fn = selective_filters.get(use_filter_name, lambda ri, m: True)
        strategy_desc = f"{use_formula_name} + {use_filter_name} @ {use_budget}kr"
    else:
        use_formula_name = top_formula_name_dyn
        use_formula_fn = top_formula_fn
        use_budget = best_budget
        use_filter_fn = lambda ri, m: True
        use_filter_name = "ALL"
        strategy_desc = f"{use_formula_name} + ALL @ {use_budget}kr"

    print(f"\n  Strategy: {strategy_desc}")

    monthly = defaultdict(lambda: {"cost": 0, "payout": 0, "hits": 0, "rounds": 0})
    all_hit_details = []

    for ri_idx, ri in enumerate(round_infos):
        m = round_conf_metrics[ri_idx]
        if not use_filter_fn(ri, m):
            continue

        row_price = ROW_PRICE.get(ri.game_type, 1.0)
        max_rows = int(use_budget / row_price)

        confs = []
        for race in ri.races:
            stats = get_race_stats(race)
            try:
                c = use_formula_fn(stats)
            except:
                c = 0
            confs.append(c)

        sim = simulate_round(ri, confs, max_rows)
        month_key = ri.round_date.strftime("%Y-%m")

        monthly[month_key]["cost"] += sim["cost"]
        monthly[month_key]["rounds"] += 1

        if sim["hit"]:
            monthly[month_key]["hits"] += 1
            monthly[month_key]["payout"] += sim["fixed_payout"]
            all_hit_details.append({
                "date": ri.round_date.isoformat(),
                "game_type": ri.game_type,
                "track": ri.track_name,
                "cost": sim["cost"],
                "payout": sim["fixed_payout"],
                "profit": sim["fixed_payout"] - sim["cost"],
                "picks": sim["picks_counts"],
                "rows": sim["total_rows"],
                "winner_strecks": [round(s, 4) for s in sim["winner_strecks"]],
            })

    print(f"\n  {'Month':>8} {'Rounds':>7} {'Hits':>5} {'Cost':>10} {'Payout':>12} {'Netto':>12} {'ROI':>8} {'Cumulative':>12}")
    print(f"  {'-'*8} {'-'*7} {'-'*5} {'-'*10} {'-'*12} {'-'*12} {'-'*8} {'-'*12}")

    cumulative_netto = 0
    monthly_data = []

    for month in sorted(monthly.keys()):
        m = monthly[month]
        netto = m["payout"] - m["cost"]
        roi = (m["payout"] - m["cost"]) / m["cost"] if m["cost"] > 0 else 0
        cumulative_netto += netto
        is_positive = "+" if netto > 0 else ""

        monthly_data.append({
            "month": month,
            "rounds": m["rounds"],
            "hits": m["hits"],
            "cost": m["cost"],
            "payout": m["payout"],
            "netto": netto,
            "roi": roi,
            "cumulative": cumulative_netto,
        })

        print(f"  {month:>8} {m['rounds']:>7} {m['hits']:>5} {m['cost']:>10,.0f} "
              f"{m['payout']:>12,.0f} {netto:>+12,.0f} {roi*100:>+7.1f}% {cumulative_netto:>+12,.0f}")

    results["monthly_breakdown"] = monthly_data

    # Show all hit details
    sep("HIT DETAILS")
    if all_hit_details:
        print(f"\n  {'Date':<12} {'Type':<5} {'Cost':>8} {'Payout':>12} {'Profit':>12} {'Picks':>30} {'Rows':>8}")
        print(f"  {'-'*12} {'-'*5} {'-'*8} {'-'*12} {'-'*12} {'-'*30} {'-'*8}")

        for hd in all_hit_details:
            picks_str = str(hd["picks"])
            print(f"  {hd['date']:<12} {hd['game_type']:<5} {hd['cost']:>8,.0f} "
                  f"{hd['payout']:>12,.0f} {hd['profit']:>+12,.0f} {picks_str:>30} {hd['rows']:>8,}")

        print(f"\n  Total hits: {len(all_hit_details)}")
        total_profit = sum(h["profit"] for h in all_hit_details)
        total_cost_played = sum(monthly[m]["cost"] for m in monthly)
        print(f"  Total profit from hits: {total_profit:+,.0f} kr")
        print(f"  Total cost (all played rounds): {total_cost_played:,.0f} kr")
    else:
        print(f"\n  No hits with this strategy.")

    results["hit_details"] = all_hit_details

    # ══════════════════════════════════════════════════════════════════════
    # STEP 7: EXPERT SIGNAL ANALYSIS
    # ══════════════════════════════════════════════════════════════════════

    sep("STEP 7: EXPERT SIGNAL ANALYSIS")

    expert_files = {
        "2026-02-14": Path("/Users/dennisdemirtok/Downloads/v85_komplett_40system.xlsx"),
        "2026-01-24": Path("/Users/dennisdemirtok/Downloads/v85_kalmar_30system.xlsx"),
        "2026-01-31": Path("/Users/dennisdemirtok/Downloads/V85_Komplett_33system.xlsx"),
    }

    expert_results = {}

    try:
        import openpyxl

        for expert_date, expert_file in expert_files.items():
            if not expert_file.exists():
                print(f"\n  Expert file not found: {expert_file}")
                continue

            print(f"\n  Parsing expert file for {expert_date}: {expert_file.name}")

            wb = openpyxl.load_workbook(expert_file, data_only=True)
            sheet = wb.active

            # Parse expert data: look for horse names and count how many systems picked each
            # The structure varies, so we try to detect it
            rows_data = []
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                       max_col=sheet.max_column, values_only=True):
                rows_data.append(list(row))

            if not rows_data:
                print(f"    Empty file")
                continue

            # Try to find race headers and horse selections
            # Look for patterns like "Avd 1", "Avd 2", etc. or "Lopp 1", "Lopp 2"
            race_headers = []
            for row_idx, row in enumerate(rows_data):
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        cell_lower = cell.lower()
                        if any(pat in cell_lower for pat in ["avd", "lopp", "race"]):
                            race_headers.append((row_idx, col_idx, cell))

            print(f"    Found {len(race_headers)} potential race headers")
            print(f"    Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} cols")

            # Parse horse names per race from expert systems
            # Strategy: count mentions of each horse across all systems
            # (Simple heuristic: look for numbered entries per race section)

            # For now, let's analyze the structure
            for row_idx in range(min(5, len(rows_data))):
                row = rows_data[row_idx]
                print(f"    Row {row_idx}: {[str(c)[:30] if c else '' for c in row[:10]]}")

            wb.close()

            # Find the matching round in our data
            target_date = date.fromisoformat(expert_date)
            matching_round = None
            for ri in round_infos:
                if ri.round_date == target_date:
                    matching_round = ri
                    break

            if matching_round:
                print(f"\n    Found matching round: {matching_round.game_type} {matching_round.round_date}")

                # Show model vs market ranking for this round
                print(f"\n    Model vs Market vs Actual for {expert_date}:")
                print(f"    {'Race':>5} {'Winner':>20} {'W.Rank(super)':>14} {'W.Streck':>10} {'Top1(super)':>20} {'Hit?':>5}")
                print(f"    {'-'*5} {'-'*20} {'-'*14} {'-'*10} {'-'*20} {'-'*5}")

                for race in matching_round.races:
                    sorted_super = sorted(race.entries, key=lambda e: e.super_score, reverse=True)
                    ranking = [e.post_position for e in sorted_super]
                    winner_entry = next((e for e in race.entries if e.post_position == race.winner_pp), None)
                    w_rank = ranking.index(race.winner_pp) + 1 if race.winner_pp in ranking else 99
                    w_name = winner_entry.horse_name if winner_entry else "?"
                    w_streck = winner_entry.bet_percentage if winner_entry else 0
                    top1_name = sorted_super[0].horse_name if sorted_super else "?"

                    in_top3 = "YES" if w_rank <= 3 else "NO"
                    print(f"    {race.race_number:>5} {w_name:>20} {w_rank:>14} {w_streck:>9.1%} {top1_name:>20} {in_top3:>5}")

                expert_results[expert_date] = {
                    "file": str(expert_file),
                    "round_date": expert_date,
                    "game_type": matching_round.game_type,
                    "note": "Expert data parsed — structure analysis above. Detailed consensus scoring would require race-specific parsing of each expert system format.",
                }
            else:
                print(f"    No matching round found for {expert_date}")

    except ImportError:
        print(f"\n  openpyxl not installed. Install with: pip install openpyxl")
        print(f"  Skipping expert analysis.")
    except Exception as ex:
        print(f"\n  Error parsing expert files: {ex}")
        import traceback
        traceback.print_exc()

    results["expert_analysis"] = expert_results

    # ══════════════════════════════════════════════════════════════════════
    # FINAL SUMMARY
    # ══════════════════════════════════════════════════════════════════════

    sep("FINAL SUMMARY & RECOMMENDATION", "=")

    print(f"\n  DATA: {len(round_infos)} rounds ({sum(1 for ri in round_infos if ri.game_type == 'V75')} V75, "
          f"{sum(1 for ri in round_infos if ri.game_type == 'V85')} V85)")

    print(f"\n  SUPER_SCORE ACCURACY (alpha=0.5):")
    print(f"    Top-1: {results['super_score_accuracy']['top1']:.1%}")
    print(f"    Top-2: {results['super_score_accuracy']['top2']:.1%}")
    print(f"    Top-3: {results['super_score_accuracy']['top3']:.1%}")

    # Best from exhaustive search
    if exhaustive_results:
        best = exhaustive_results[0]
        print(f"\n  BEST STRATEGY (from {len(exhaustive_results)} tested):")
        print(f"    Confidence formula: {best['formula']}")
        print(f"    Selective filter:   {best['filter']}")
        print(f"    Budget per round:   {best['budget']} kr")
        print(f"    Rounds played:      {best['played']}/{len(round_infos)}")
        print(f"    Hits:               {best['hits']}")
        print(f"    Hit rate:           {best['hit_rate']:.1%}")
        print(f"    Total cost:         {best['total_cost']:,.0f} kr")
        print(f"    Total payout:       {best['total_payout']:,.0f} kr")
        print(f"    Netto:              {best['netto']:+,.0f} kr")
        print(f"    ROI:                {best['roi']*100:+.1f}%")

        if best['hits'] > 0 and best['hit_list']:
            print(f"\n    Hit details:")
            for hd in best['hit_list']:
                profit = hd['payout'] - hd['cost']
                print(f"      {hd['date']} {hd['game_type']}: cost={hd['cost']:,.0f}, "
                      f"payout={hd['payout']:,.0f}, profit={profit:+,.0f}, picks={hd['picks']}")

        # Positive ROI strategies
        positive = [r for r in exhaustive_results if r["roi"] > 0 and r["played"] >= 5]
        print(f"\n  POSITIVE ROI STRATEGIES (min 5 rounds): {len(positive)}")
        for i, p in enumerate(positive[:10]):
            print(f"    {i+1}. [{p['formula']}] + [{p['filter']}] @ {p['budget']}kr: "
                  f"ROI={p['roi']*100:+.1f}%, {p['hits']} hits/{p['played']} played, "
                  f"netto={p['netto']:+,.0f} kr")

    # Best dynamic strategy
    if best_dynamic.get("roi", -999) > 0:
        bd = best_dynamic
        print(f"\n  BEST DYNAMIC BUDGET STRATEGY:")
        print(f"    Strategy:    {bd['strategy']}")
        print(f"    ROI:         {bd['roi']*100:+.1f}%")
        print(f"    Netto:       {bd['data']['netto']:+,.0f} kr")

    # Monthly profitable analysis
    if monthly_data:
        profitable_months = [m for m in monthly_data if m["netto"] > 0]
        losing_months = [m for m in monthly_data if m["netto"] < 0]
        print(f"\n  MONTHLY STATS:")
        print(f"    Profitable months: {len(profitable_months)}/{len(monthly_data)}")
        print(f"    Losing months:     {len(losing_months)}/{len(monthly_data)}")
        if profitable_months:
            print(f"    Best month:        {max(monthly_data, key=lambda m: m['netto'])['month']} "
                  f"({max(monthly_data, key=lambda m: m['netto'])['netto']:+,.0f} kr)")
        if losing_months:
            print(f"    Worst month:       {min(monthly_data, key=lambda m: m['netto'])['month']} "
                  f"({min(monthly_data, key=lambda m: m['netto'])['netto']:+,.0f} kr)")

    # THE VERDICT
    sep("THE VERDICT: CAN WE ACHIEVE POSITIVE ROI?", "=")

    any_positive = any(r["roi"] > 0 for r in exhaustive_results[:30]) if exhaustive_results else False
    any_positive_5plus = any(r["roi"] > 0 and r["played"] >= 5 for r in exhaustive_results) if exhaustive_results else False

    if any_positive_5plus:
        print(f"\n  YES — Positive ROI is achievable with the right strategy.")
        best_viable = next(r for r in exhaustive_results if r["roi"] > 0 and r["played"] >= 5)
        print(f"\n  TOP RECOMMENDATION:")
        print(f"    Formula:        {best_viable['formula']}")
        print(f"    Filter:         {best_viable['filter']}")
        print(f"    Budget:         {best_viable['budget']} kr/round")
        print(f"    Expected:       {best_viable['hits']} hits on {best_viable['played']} rounds = {best_viable['hit_rate']:.1%}")
        print(f"    Expected ROI:   {best_viable['roi']*100:+.1f}%")
        print(f"    Expected netto: {best_viable['netto']:+,.0f} kr per {best_viable['played']} rounds")

        print(f"\n  CAVEATS:")
        print(f"    - This is backtested on {len(round_infos)} rounds, not live trading")
        print(f"    - Overfitting risk: many formulas tested, best may not generalize")
        print(f"    - Estimated payouts are approximate (actual ATG payouts may differ)")
        print(f"    - Small sample size for selective filters")
        print(f"    - Expert signal not yet integrated (could improve further)")
    elif any_positive:
        print(f"\n  PARTIALLY — Some configurations show positive ROI but on few rounds.")
        print(f"    More data needed to confirm these are real signals vs noise.")
    else:
        print(f"\n  NOT YET — No configuration achieves positive ROI with current model accuracy.")
        print(f"\n  What would be needed:")
        avg_cost = statistics.mean([r["total_cost"]/r["played"] for r in exhaustive_results[:5] if r["played"] > 0])
        print(f"    Current avg cost/round: {avg_cost:,.0f} kr")
        print(f"    Current per-race top-3 rate: {results['super_score_accuracy']['top3']:.1%}")

        for target in [0.70, 0.75, 0.80]:
            p_all_7 = target ** 7
            p_all_8 = target ** 8
            print(f"    If top-3 = {target:.0%}: P(hit V75) = {p_all_7:.1%}, P(hit V85) = {p_all_8:.1%}")

    return results


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

async def main():
    start_time = time.time()

    print("=" * 110)
    print("  FINAL STRATEGY ANALYSIS — V75/V85 Profitability with Partial Gardering")
    print("  Budget: 3000-4000 kr/round | Super score: 0.5*model + 0.5*market")
    print("  Testing 25 confidence formulas x 20 selective filters x 5 budget levels")
    print("=" * 110)

    # Load data
    logger.info("Loading all cached V75+V85 rounds...")
    all_rounds = await load_rounds()
    logger.info(f"Loaded {len(all_rounds)} rounds in {time.time()-start_time:.0f}s")

    # Analyze
    logger.info("Analyzing all rounds with CompositeAnalyzer...")
    round_infos = analyze_all_rounds(all_rounds)

    v75 = [r for r in round_infos if r.game_type == "V75"]
    v85 = [r for r in round_infos if r.game_type == "V85"]
    print(f"\n  Rounds: {len(round_infos)} total ({len(v75)} V75, {len(v85)} V85)")
    dates = [r.round_date for r in round_infos]
    if dates:
        print(f"  Period: {min(dates)} to {max(dates)}")

    # Run full analysis
    results = run_full_analysis(round_infos)

    # Save results
    output = {
        "metadata": {
            "num_rounds": len(round_infos),
            "v75_rounds": len(v75),
            "v85_rounds": len(v85),
            "period_start": min(dates).isoformat() if dates else "",
            "period_end": max(dates).isoformat() if dates else "",
            "total_races": sum(len(r.races) for r in round_infos),
            "elapsed_seconds": round(time.time() - start_time, 1),
        },
        "super_score_accuracy": results.get("super_score_accuracy", {}),
        "best_formula": {
            "formula": results.get("best_formula", {}).get("formula", ""),
            "budget": results.get("best_formula", {}).get("budget", 0),
            "roi": results.get("best_formula", {}).get("roi", 0),
        },
        "formula_ranking": results.get("formula_ranking", []),
        "selective_results": results.get("selective_results", {}),
        "best_selective": {
            k: v for k, v in results.get("best_selective", {}).items()
            if k != "data"
        },
        "exhaustive_top30": [
            {k: v for k, v in r.items() if k != "hit_list"}
            for r in results.get("exhaustive_top30", [])
        ],
        "dynamic_results": results.get("dynamic_results", {}),
        "best_dynamic": {
            k: v for k, v in results.get("best_dynamic", {}).items()
            if k != "data"
        },
        "monthly_breakdown": results.get("monthly_breakdown", []),
        "hit_details": results.get("hit_details", []),
        "expert_analysis": results.get("expert_analysis", {}),
    }

    output_path = Path(__file__).parent / "final_strategy.json"
    output_path.write_text(json.dumps(output, indent=2, ensure_ascii=False, default=str))
    logger.info(f"Results saved to {output_path}")

    elapsed = time.time() - start_time
    print(f"\n  Total analysis time: {elapsed:.0f}s ({elapsed/60:.1f} min)")
    logger.info(f"Done in {elapsed:.0f}s")


if __name__ == "__main__":
    asyncio.run(main())
