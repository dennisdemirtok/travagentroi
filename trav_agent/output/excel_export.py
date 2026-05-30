"""Excel-export av betting-system / rekommendationer.

Tar en ``SystemPlan`` (från system_builder) + en ``GameRound`` och skapar en
.xlsx-fil med hästnamn, breddning, svårighet och motivering per avdelning, samt
en kompakt "spelrad" som kan klistras in på ATG.

Användning::

    from trav_agent.output.excel_export import system_to_xlsx_bytes
    blob = system_to_xlsx_bytes(plan, game_round)
    open("system.xlsx", "wb").write(blob)
"""

from __future__ import annotations

import io
from typing import Optional

# ── Färgtema (matchar dashboardens A/B/C/D) ─────────────────────────────────
_C_HEADER = "1E293B"      # mörk slate
_C_HEADER_TXT = "FFFFFF"
_C_SPIK = "DCFCE7"        # grön (spik)
_C_KORT = "DBEAFE"        # blå (kort)
_C_MEDEL = "FEF3C7"       # amber (medel)
_C_BRED = "FFEDD5"        # orange (bred)
_C_META = "F3F4F6"        # ljusgrå metarad
_C_ACCENT = "F59E0B"

_LEG_TYPE_LABEL = {
    "spik": "SPIK", "kort": "Kort", "medel": "Medel", "bred": "Bred",
}
_LEG_TYPE_COLOR = {
    "spik": _C_SPIK, "kort": _C_KORT, "medel": _C_MEDEL, "bred": _C_BRED,
}


def _horse_name(game_round, race_number: int, post: int) -> str:
    """Slå upp hästnamn för (avd, spår)."""
    race = game_round.get_race(race_number) if game_round else None
    if not race:
        return f"#{post}"
    for e in race.entries:
        if e.post_position == post:
            if e.horse and e.horse.name:
                return e.horse.name
    return f"#{post}"


def build_system_workbook(plan, game_round=None, *, title: Optional[str] = None):
    """Bygg en openpyxl Workbook för en SystemPlan.

    Returns:
        openpyxl.Workbook
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "System"

    bold = Font(bold=True, color="1E293B")
    white_bold = Font(bold=True, color=_C_HEADER_TXT, size=11)
    accent_bold = Font(bold=True, color=_C_ACCENT, size=14)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")

    gt = plan.game_type
    rd = plan.round_date
    legs = sorted(plan.legs, key=lambda l: l.race_number)

    # ── Titelrad ──
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = title or f"{gt} {rd} — Systemspel"
    c.font = accent_bold
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Metadata-block ──
    meta = [
        ("Strategi", plan.strategy_name or "Smart ABCD"),
        ("Budget", f"{plan.budget:.0f} kr"),
        ("Rader", f"{plan.total_rows}"),
        ("Kostnad", f"{plan.total_cost:.0f} kr"),
        ("Spikar", f"{plan.num_spikes}"),
        ("P(alla rätt)", f"{plan.predicted_hit_prob:.1%}" if plan.predicted_hit_prob else "—"),
    ]
    row = 2
    for label, val in meta:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = bold
        lc.fill = PatternFill("solid", fgColor=_C_META)
        vc = ws.cell(row=row, column=2, value=val)
        vc.font = Font(color="1E293B")
        vc.fill = PatternFill("solid", fgColor=_C_META)
        row += 1

    # ── Tabellhuvud ──
    header_row = row + 1
    headers = ["Avd", "Lopp", "Typ", "Antal", "Hästar (spår — namn)", "Svårighet"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = white_bold
        cell.fill = PatternFill("solid", fgColor=_C_HEADER)
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[header_row].height = 20

    # ── Datarader ──
    r = header_row + 1
    for leg in legs:
        race = game_round.get_race(leg.race_number) if game_round else None
        dist = f"{race.distance}m {race.start_method.value}" if race else ""
        type_label = _LEG_TYPE_LABEL.get(leg.leg_type, leg.leg_type)
        fill = PatternFill("solid", fgColor=_LEG_TYPE_COLOR.get(leg.leg_type, "FFFFFF"))

        horses = " · ".join(
            f"{p} {_horse_name(game_round, leg.race_number, p)}"
            for p in leg.picks[: leg.num_picks]
        )

        vals = [
            leg.race_number,
            dist,
            type_label,
            leg.num_picks,
            horses,
            round(leg.difficulty, 0),
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.border = border
            cell.fill = fill
            if col in (1, 3, 4, 6):
                cell.alignment = center
            else:
                cell.alignment = wrap
            if col == 3:
                cell.font = bold
        r += 1

    # ── Kolumnbredder ──
    widths = [6, 18, 8, 7, 60, 11]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Blad 2: Kompakt spelrad (för ATG-inmatning) ──
    ws2 = wb.create_sheet("Spelrad")
    ws2["A1"] = f"{gt} {rd} — kompakt spelrad"
    ws2["A1"].font = accent_bold
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 50
    rr = 3
    ws2.cell(row=rr, column=1, value="Avd").font = white_bold
    ws2.cell(row=rr, column=1).fill = PatternFill("solid", fgColor=_C_HEADER)
    ws2.cell(row=rr, column=2, value="Hästar").font = white_bold
    ws2.cell(row=rr, column=2).fill = PatternFill("solid", fgColor=_C_HEADER)
    rr += 1
    for leg in legs:
        nums = ",".join(str(p) for p in leg.picks[: leg.num_picks])
        ws2.cell(row=rr, column=1, value=leg.race_number).alignment = center
        ws2.cell(row=rr, column=2, value=nums)
        if leg.leg_type == "spik":
            for col in (1, 2):
                ws2.cell(row=rr, column=col).fill = PatternFill("solid", fgColor=_C_SPIK)
        rr += 1

    return wb


def system_to_xlsx_bytes(plan, game_round=None, *, title: Optional[str] = None) -> bytes:
    """Bygg systemet och returnera .xlsx som bytes (för HTTP-respons)."""
    wb = build_system_workbook(plan, game_round, title=title)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_system_xlsx(plan, game_round, path: str, *, title: Optional[str] = None) -> str:
    """Spara systemet till en .xlsx-fil på disk. Returnerar sökvägen."""
    wb = build_system_workbook(plan, game_round, title=title)
    wb.save(path)
    return path
